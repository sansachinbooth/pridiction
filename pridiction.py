#!/usr/bin/env python3
"""
pridict.py - Enhanced Crypto Prediction Tool with Accuracy Checking
New feature: Compare predictions with live data to measure accuracy
"""

import os
import sys
import time
import math
import json
import argparse
from datetime import datetime, timedelta
from typing import Tuple, Optional, Dict, Any

import requests
import numpy as np
import pandas as pd

from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score

# color output
try:
    from colorama import init as colorama_init, Fore, Style, Back
    colorama_init(autoreset=True)
except Exception:
    # fallback no color
    class _C:
        def __getattr__(self, name): return ""
    Fore = Style = Back = _C()

# Optional imports for enhanced Excel reporting
try:
    import openpyxl
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import load_workbook
    EXCEL_ENHANCED = True
except Exception:
    EXCEL_ENHANCED = False

# Optional imports
try:
    import ollama  # local ollama client if installed
    OLLAMA_AVAILABLE = True
except Exception:
    OLLAMA_AVAILABLE = False

# Constants
COINGECKO_API = "https://api.coingecko.com/api/v3"
BINANCE_API = "https://api.binance.com/api/v3"

# ----- Output formatting functions -----

def print_header(title: str, level: int = 1):
    if level == 1:
        print(Fore.CYAN + "=" * 70)
        print(Fore.CYAN + title.center(70))
        print(Fore.CYAN + "=" * 70)
    elif level == 2:
        print(Fore.GREEN + "─" * 50)
        print(Fore.GREEN + f"▶ {title}")
        print(Fore.GREEN + "─" * 50)
    else:
        print(Fore.YELLOW + f"● {title}")

def print_bullet(text: str, color=Fore.WHITE, indent: int = 2):
    indent_str = " " * indent
    print(f"{indent_str}{color}• {text}")

def print_status(text: str, status: str = "info"):
    colors = {
        "info": Fore.BLUE,
        "success": Fore.GREEN,
        "warning": Fore.YELLOW,
        "error": Fore.RED
    }
    color = colors.get(status, Fore.WHITE)
    print(f"{color}↳ {text}")

# ----- Utilities and indicators -----

def ema(series: pd.Series, period: int) -> pd.Series:
    """Calculate Exponential Moving Average with proper handling"""
    if len(series) < period:
        # Not enough data, return series of NaN
        return pd.Series([float('nan')] * len(series), index=series.index)
    return series.ewm(span=period, adjust=False).mean()

def rsi(series: pd.Series, period: int = 14) -> pd.Series:
    """Calculate RSI with proper error handling"""
    if len(series) < period + 1:
        return pd.Series([50.0] * len(series), index=series.index)  # Neutral RSI
    
    delta = series.diff()
    up = delta.clip(lower=0)
    down = -1 * delta.clip(upper=0)
    
    # Use SMA for initial values then EMA
    ma_up = up.rolling(window=period, min_periods=period).mean()
    ma_down = down.rolling(window=period, min_periods=period).mean()
    
    # Handle division by zero
    rs = ma_up / ma_down
    rsi = 100 - (100 / (1 + rs))
    
    return rsi.fillna(50.0)  # Fill NaN with neutral 50

def macd(series: pd.Series, fast: int = 12, slow: int = 26, signal: int = 9) -> Tuple[pd.Series, pd.Series, pd.Series]:
    """Calculate MACD with proper validation"""
    if len(series) < slow:
        nan_series = pd.Series([float('nan')] * len(series), index=series.index)
        return nan_series, nan_series, nan_series
    
    fast_ema = ema(series, fast)
    slow_ema = ema(series, slow)
    macd_line = fast_ema - slow_ema
    signal_line = ema(macd_line, signal)
    hist = macd_line - signal_line
    return macd_line, signal_line, hist

def atr(df: pd.DataFrame, period: int = 14) -> pd.Series:
    """Calculate Average True Range"""
    if len(df) < period:
        return pd.Series([float('nan')] * len(df), index=df.index)
    
    high_low = df['high'] - df['low']
    high_close = (df['high'] - df['close'].shift()).abs()
    low_close = (df['low'] - df['close'].shift()).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    return tr.rolling(window=period, min_periods=1).mean()

def bollinger_bands(series: pd.Series, period: int = 20, std: int = 2) -> Tuple[pd.Series, pd.Series, pd.Series]:
    """Calculate Bollinger Bands"""
    if len(series) < period:
        nan_series = pd.Series([float('nan')] * len(series), index=series.index)
        return nan_series, nan_series, nan_series
    
    sma = series.rolling(window=period, min_periods=1).mean()
    rolling_std = series.rolling(window=period, min_periods=1).std()
    upper_band = sma + (rolling_std * std)
    lower_band = sma - (rolling_std * std)
    return upper_band, sma, lower_band

def is_valid_value(x) -> bool:
    """Check if a value is valid (not None, NaN, or Inf)"""
    try:
        return x is not None and not (math.isnan(x) or math.isinf(x))
    except Exception:
        return False

# ----- Data fetchers -----

def fetch_coin_gecko_market_chart(coin_id: str, vs_currency: str, days: float) -> pd.DataFrame:
    """
    Fetch coin market chart (prices, market_caps, total_volumes).
    - days can be fractional (e.g., 0.0208 for 30 minutes) but CoinGecko may restrict minute-resolution to 1 day.
    Returns DataFrame with columns: timestamp, price, market_cap, total_volume
    """
    url = f"{COINGECKO_API}/coins/{coin_id}/market_chart"
    params = {"vs_currency": vs_currency, "days": days}
    r = requests.get(url, params=params, timeout=20)
    r.raise_for_status()
    data = r.json()
    prices = data.get("prices", [])
    # prices: [ [timestamp_ms, price], ... ]
    df = pd.DataFrame(prices, columns=['timestamp_ms', 'price'])
    df['timestamp'] = pd.to_datetime(df['timestamp_ms'], unit='ms')
    df = df[['timestamp', 'price']]
    return df

def coingecko_ohlcv_from_prices(df_prices: pd.DataFrame, timeframe: str) -> pd.DataFrame:
    """
    Convert raw price points to OHLCV by resampling.
    timeframe examples: '10m', '20m', '30m', '1h', '4h', '1d', '1y'
    """
    rule_map = {
        '10m': '10min', '20m': '20min', '30m': '30min',
        '1h': '1H', '4h': '4H', '1d': '1D', '1y': '1Y'
    }
    if timeframe not in rule_map:
        raise ValueError("Unsupported timeframe for resample")
    rule = rule_map[timeframe]
    df = df_prices.set_index('timestamp').resample(rule).agg({'price': ['first', 'max', 'min', 'last', 'count']})
    df.columns = ['open', 'high', 'low', 'close', 'volume']
    df = df.dropna().reset_index()
    return df

def fetch_binance_klines(symbol: str, interval: str, limit: int = 500) -> pd.DataFrame:
    """
    Fetch klines from Binance: returns DataFrame with timestamp, open, high, low, close, volume
    interval examples: '1m', '5m', '15m', '1h', '4h', '1d'
    """
    url = f"{BINANCE_API}/klines"
    params = {"symbol": symbol.upper(), "interval": interval, "limit": limit}
    r = requests.get(url, params=params, timeout=1500)
    r.raise_for_status()
    data = r.json()
    cols = ['open_time', 'open', 'high', 'low', 'close', 'volume', 'close_time', 'qav', 'num_trades',
            'taker_base_vol', 'taker_quote_vol', 'ignore']
    df = pd.DataFrame(data, columns=cols)
    df['timestamp'] = pd.to_datetime(df['open_time'], unit='ms')
    for c in ['open', 'high', 'low', 'close', 'volume']:
        df[c] = df[c].astype(float)
    df = df[['timestamp', 'open', 'high', 'low', 'close', 'volume']]
    return df

# ----- Feature engineering -----

def add_technical_indicators(df: pd.DataFrame) -> pd.DataFrame:
    """Enhanced version with proper error handling and data validation"""
    if df.empty:
        print_status("Empty dataframe provided to technical indicators", "warning")
        return df
    
    df = df.copy()
    
    # Ensure numeric types with proper error handling
    for col in ['open', 'high', 'low', 'close', 'volume']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Remove any rows with invalid price data
    df = df.dropna(subset=['close', 'high', 'low', 'open'])
    
    if df.empty:
        print_status("No valid price data after cleaning", "error")
        return df
    
    n_rows = len(df)
    print_status(f"Processing {n_rows} rows for technical indicators", "info")
    
    try:
        # Moving averages with minimum period check
        df['sma20'] = df['close'].rolling(window=min(20, n_rows), min_periods=1).mean()
        df['sma50'] = df['close'].rolling(window=min(50, n_rows), min_periods=1).mean()
        
        # EMAs
        df['ema8'] = ema(df['close'], min(8, n_rows))
        df['ema21'] = ema(df['close'], min(21, n_rows))
        df['ema50'] = ema(df['close'], min(50, n_rows))
        
        # RSI with safe calculation
        df['rsi14'] = rsi(df['close'], min(14, n_rows))
        
        # MACD
        macd_line, signal_line, hist_line = macd(df['close'])
        df['macd'] = macd_line
        df['macd_signal'] = signal_line
        df['macd_hist'] = hist_line
        
        # ATR
        df['atr14'] = atr(df, min(14, n_rows))
        
        # Bollinger Bands
        upper_bb, middle_bb, lower_bb = bollinger_bands(df['close'])
        df['bb_upper'] = upper_bb
        df['bb_middle'] = middle_bb
        df['bb_lower'] = lower_bb
        df['bb_width'] = (upper_bb - lower_bb) / middle_bb  # Bollinger Band Width
        
        # Volume indicators
        df['volume_sma20'] = df['volume'].rolling(window=min(20, n_rows), min_periods=1).mean()
        df['volume_ratio'] = df['volume'] / df['volume_sma20']
        df['volume_ratio'] = df['volume_ratio'].replace([np.inf, -np.inf], np.nan)
        
        # Price changes
        df['ret1'] = df['close'].pct_change(1)
        df['ret3'] = df['close'].pct_change(3)
        df['ret7'] = df['close'].pct_change(7)
        
        # Support/Resistance levels (simplified)
        df['resistance'] = df['high'].rolling(window=min(20, n_rows), min_periods=1).max()
        df['support'] = df['low'].rolling(window=min(20, n_rows), min_periods=1).min()
        
        # Fix: Replace deprecated fillna with method with ffill() and bfill()
        for col in df.columns:
            if df[col].dtype in ['float64', 'int64']:
                if 'rsi' in col:
                    df[col] = df[col].fillna(50.0)  # Neutral RSI
                elif 'macd' in col or 'bb' in col:
                    df[col] = df[col].fillna(0.0)
                else:
                    # Use ffill() and bfill() instead of deprecated fillna with method
                    df[col] = df[col].ffill().bfill().fillna(0.0)
        
        # Final cleanup - remove rows that are still mostly NaN
        df = df.dropna(thresh=len(df.columns) // 2)
        
        if not df.empty:
            print_status(f"Technical indicators calculated successfully. Final dataset: {len(df)} rows", "success")
        else:
            print_status("No valid data after technical indicator calculation", "warning")
            
        return df.reset_index(drop=True)
        
    except Exception as e:
        print_status(f"Error in technical indicator calculation: {str(e)}", "error")
        # Return original dataframe if technical indicators fail
        return df

# ----- ML model -----

def train_ml_predictor(df: pd.DataFrame, feature_cols: list, target_col: str = 'close', test_size: float = 0.15):
    """
    Train a RandomForestRegressor to predict next candle close.
    Returns (model, X_test, y_test, preds)
    """
    df = df.copy()
    
    # Ensure we have enough data
    if len(df) < 50:
        print_status("Not enough data for ML training (min 50 rows required)", "warning")
        return None
        
    # target is next candle close
    df['target'] = df[target_col].shift(-1)
    df = df.dropna()
    
    if df.empty or len(df) < 30:
        print_status("Not enough valid data for ML training after preprocessing", "warning")
        return None
        
    # Ensure all feature columns exist and are numeric
    available_features = [col for col in feature_cols if col in df.columns]
    missing_features = set(feature_cols) - set(available_features)
    
    if missing_features:
        print_status(f"Missing features: {missing_features}. Using available features: {available_features}", "warning")
    
    if not available_features:
        print_status("No valid features available for ML training", "error")
        return None
    
    X = df[available_features]
    y = df['target']
    
    # Remove any infinite values
    X = X.replace([np.inf, -np.inf], np.nan)
    X = X.fillna(X.mean())
    
    if X.isnull().any().any():
        print_status("NaN values detected in features after cleaning", "warning")
        X = X.fillna(0)
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, shuffle=False)
    
    if len(X_train) == 0:
        print_status("No training data available after split", "error")
        return None
        
    try:
        model = RandomForestRegressor(n_estimators=100, random_state=42, n_jobs=-1, max_depth=10)
        model.fit(X_train, y_train)
        preds = model.predict(X_test)
        mae = mean_absolute_error(y_test, preds)
        
        print_status(f"ML model trained successfully. MAE: {mae:.6f}", "success")
        return {
            'model': model,
            'X_test': X_test,
            'y_test': y_test,
            'preds': preds,
            'mae': mae,
            'last_row': df.iloc[-1] if not df.empty else None,
            'feature_cols': available_features
        }
    except Exception as e:
        print_status(f"ML training failed: {str(e)}", "error")
        return None

def predict_future_prices(model, last_data: pd.Series, feature_cols: list, periods: int = 365) -> pd.DataFrame:
    """
    Predict future prices for multiple periods (for 1-year prediction)
    """
    predictions = []
    current_data = last_data.copy()
    
    for i in range(periods):
        if i > 0:
            # Update the current_data with the previous prediction for recursive forecasting
            current_data['close'] = pred_price
            # Recalculate some basic indicators based on the new close
            if 'ema8' in feature_cols:
                current_data['ema8'] = (current_data['ema8'] * 7 + pred_price) / 8  # Simplified EMA update
        
        # Prepare features for prediction
        try:
            features = current_data[feature_cols].values.reshape(1, -1)
            pred_price = float(model.predict(features)[0])
        except Exception as e:
            print_status(f"Prediction error at step {i}: {e}", "warning")
            pred_price = current_data['close']  # Fallback to current price
            
        predictions.append({
            'period': i + 1,
            'predicted_price': pred_price,
            'date': datetime.now() + timedelta(days=i)
        })
    
    return pd.DataFrame(predictions)

# ----- News sentiment (simple) -----

def fetch_news_sentiment(query: str, api_key: Optional[str], page_size: int = 5) -> Dict[str, Any]:
    """
    Uses NewsAPI if api_key provided. Returns simple sentiment (pos/neg/neutral) based on naive word list.
    """
    if not api_key:
        return {'available': False, 'reason': 'No NEWSAPI_KEY provided'}
    url = "https://newsapi.org/v2/everything"
    params = {'q': query, 'language': 'en', 'pageSize': page_size, 'sortBy': 'publishedAt'}
    headers = {'Authorization': api_key}
    try:
        r = requests.get(url, params=params, headers=headers, timeout=1500)
        if r.status_code != 200:
            return {'available': False, 'reason': f"newsapi error {r.status_code}: {r.text}"}
        data = r.json()
        articles = data.get('articles', [])
        # very naive sentiment
        pos_words = {'gain', 'surge', 'bull', 'rise', 'record', 'beat', 'positive', 'up'}
        neg_words = {'drop', 'crash', 'bear', 'decline', 'fall', 'hack', 'negative', 'down', 'dump'}
        scores = []
        for a in articles:
            text = (a.get('title') or '') + ' ' + (a.get('description') or '')
            t = text.lower()
            score = sum(1 for w in pos_words if w in t) - sum(1 for w in neg_words if w in t)
            scores.append({'title': a.get('title'), 'publishedAt': a.get('publishedAt'), 'score': score, 'url': a.get('url')})
        avg = float(np.mean([s['score'] for s in scores])) if scores else 0.0
        return {'available': True, 'avg_score': avg, 'articles': scores}
    except Exception as e:
        return {'available': False, 'reason': f"Exception: {str(e)}"}

# ----- Whale check (basic using Binance recent trades) -----

def fetch_binance_recent_trades(symbol: str, limit: int = 100) -> pd.DataFrame:
    try:
        url = f"{BINANCE_API}/aggTrades"
        params = {"symbol": symbol.upper(), "limit": limit}
        r = requests.get(url, params=params, timeout=10)
        r.raise_for_status()
        trades = r.json()
        # trades: [{a,b,p,q,f,l,T,m,...}]
        df = pd.DataFrame(trades)
        if df.empty:
            return df
        df['price'] = df['p'].astype(float)
        df['qty'] = df['q'].astype(float)
        df['timestamp'] = pd.to_datetime(df['T'], unit='ms')
        return df[['timestamp', 'price', 'qty']]
    except Exception as e:
        print(f"Error fetching Binance trades: {e}")
        return pd.DataFrame()

def analyze_whales_from_trades(trades_df: pd.DataFrame) -> Dict[str, Any]:
    if trades_df is None or trades_df.empty:
        return {'available': False, 'reason': 'no_trades'}
    q = trades_df['qty']
    median = q.median()
    threshold = median * 10  # heuristic: 10x median quantity is considered a whale
    whales = trades_df[trades_df['qty'] >= threshold]
    return {'available': True, 'median_qty': float(median), 'threshold': float(threshold), 'whales': whales.to_dict(orient='records')}

# ----- Ollama integration (optional) -----

def generate_ollama_summary(prompt: str, host: str = "http://localhost", port: int = 11500, model: str = "gemma3:4b") -> Dict[str, Any]:
    """Fixed URL construction for Ollama API"""
    try:
        # Fix URL construction - remove duplicate port
        base_url = host
        if not base_url.startswith(('http://', 'https://')):
            base_url = f"http://{base_url}"
        
        # Clean up port formatting
        if ':' in base_url and base_url.rfind(':') > 4:  # If port already in host
            url = f"{base_url}/api/generate"
        else:
            url = f"{base_url}:{port}/api/generate"
        
        payload = {"model": model, "prompt": prompt, "stream": False}
        r = requests.post(url, json=payload, timeout=3000)
        r.raise_for_status()
        resp = r.json()
        return {'ok': True, 'text': resp.get('response', '')}
    except requests.exceptions.ConnectionError:
        return {'ok': False, 'error': 'Ollama not running - start with: ollama serve'}
    except Exception as e:
        return {'ok': False, 'error': f'Ollama error: {str(e)}'}

# ----- Trading recommendation -----

def recommend_trade(last_price: float, predicted_price: float, atr_val: float, rsi_val: float, timeframe: str = '1h') -> Dict[str, Any]:
    """
    Enhanced recommendation with RSI consideration and timeframe-based sale time prediction
    """
    diff = predicted_price - last_price if predicted_price is not None and is_valid_value(predicted_price) else 0
    diff_pct = (diff / last_price) * 100.0 if predicted_price is not None and is_valid_value(predicted_price) else 0
    
    # Multi-factor decision
    factors = []
    
    # Price prediction factor
    if diff_pct > 1.0:
        factors.append(("price_prediction", "strong_bullish", 2))
    elif diff_pct > 0.3:
        factors.append(("price_prediction", "bullish", 1))
    elif diff_pct < -1.0:
        factors.append(("price_prediction", "strong_bearish", -2))
    elif diff_pct < -0.3:
        factors.append(("price_prediction", "bearish", -1))
    else:
        factors.append(("price_prediction", "neutral", 0))
    
    # RSI factor
    if rsi_val > 70:
        factors.append(("rsi", "overbought", -1))
    elif rsi_val < 30:
        factors.append(("rsi", "oversold", 1))
    else:
        factors.append(("rsi", "neutral", 0))
    
    # Calculate weighted score
    total_score = sum(score for _, _, score in factors)
    
    # Determine advice
    if total_score >= 2:
        advice = 'STRONG BUY'
        confidence = 'High'
    elif total_score >= 1:
        advice = 'BUY'
        confidence = 'Medium'
    elif total_score <= -2:
        advice = 'STRONG SELL'
        confidence = 'High'
    elif total_score <= -1:
        advice = 'SELL'
        confidence = 'Medium'
    else:
        advice = 'HOLD'
        confidence = 'Low'
    
    # Risk management
    stoploss = last_price - (2 * atr_val) if is_valid_value(atr_val) and atr_val > 0 else last_price * 0.98  # 2% stoploss fallback
    take_profit = predicted_price if predicted_price and is_valid_value(predicted_price) else last_price * 1.02  # 2% target if no prediction
    
    risk_reward = (take_profit - last_price) / (last_price - stoploss) if stoploss and stoploss < last_price else None
    
    # Sale time prediction based on timeframe and recommendation
    current_time = datetime.now()
    sale_time = None
    
    if advice in ['BUY', 'STRONG BUY']:
        # Predict sale time based on timeframe
        if timeframe == '10m':
            sale_time = current_time + timedelta(minutes=30)  # 30 minutes for 10m timeframe
        elif timeframe == '20m':
            sale_time = current_time + timedelta(minutes=60)  # 1 hour for 20m timeframe
        elif timeframe == '30m':
            sale_time = current_time + timedelta(hours=2)  # 2 hours for 30m timeframe
        elif timeframe == '1h':
            sale_time = current_time + timedelta(hours=4)  # 4 hours for 1h timeframe
        elif timeframe == '4h':
            sale_time = current_time + timedelta(days=1)  # 1 day for 4h timeframe
        elif timeframe == '1d':
            sale_time = current_time + timedelta(days=3)  # 3 days for 1d timeframe
        elif timeframe == '1y':
            sale_time = current_time + timedelta(days=365)  # 1 year for 1y timeframe
        else:
            sale_time = current_time + timedelta(hours=24)  # Default 24 hours
    
    return {
        'advice': advice,
        'confidence': confidence,
        'diff_pct': diff_pct,
        'factors': factors,
        'total_score': total_score,
        'stoploss': stoploss,
        'take_profit': take_profit,
        'risk_reward_ratio': risk_reward,
        'sale_time': sale_time
    }

def get_rsi_strength(rsi_value: float) -> Tuple[str, str]:
    """Get RSI strength description"""
    if not is_valid_value(rsi_value):
        return "UNKNOWN", Fore.WHITE
    if rsi_value >= 70:
        return "OVERBOUGHT", Fore.RED
    elif rsi_value >= 60:
        return "BULLISH", Fore.YELLOW
    elif rsi_value >= 40:
        return "NEUTRAL", Fore.WHITE
    elif rsi_value >= 30:
        return "BEARISH", Fore.YELLOW
    else:
        return "OVERSOLD", Fore.GREEN

def get_trend_strength(ema_short: float, ema_long: float, price: float) -> Tuple[str, str]:
    """Determine trend strength based on EMAs and price"""
    if not all(is_valid_value(x) for x in [ema_short, ema_long, price]):
        return "UNKNOWN", Fore.WHITE
        
    if price > ema_short > ema_long:
        return "STRONG UPTREND", Fore.GREEN
    elif ema_short > price > ema_long:
        return "UPTREND", Fore.YELLOW
    elif price < ema_short < ema_long:
        return "STRONG DOWNTREND", Fore.RED
    elif ema_short < price < ema_long:
        return "DOWNTREND", Fore.YELLOW
    else:
        return "SIDEWAYS", Fore.WHITE

# ----- Simplified Excel Export -----

def create_simplified_export(coin: str, symbol: str, timeframe: str, last_price: float, 
                           predicted_price: float, recommendation: Dict, rsi_value: float, 
                           output_path: str, timestamp_str: str = None):
    """
    Create a simplified Excel export with only the required columns:
    - Sl No., Symbol, Crypto Name, Timestamp, Timeframe, Current Price, 
    - Prediction Value, Signal, Stoploss, RSI, Buy Time, Sale Time
    """
    if timestamp_str is None:
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create the simplified data
    current_time = datetime.now()
    
    # Determine buy/sale time based on signal
    buy_time = current_time if recommendation.get('advice') in ['BUY', 'STRONG BUY'] else None
    sale_time = recommendation.get('sale_time') if recommendation.get('advice') in ['BUY', 'STRONG BUY'] else current_time if recommendation.get('advice') in ['SELL', 'STRONG SELL'] else None
    
    data = {
        'Sl No': [1],
        'Symbol': [symbol.upper() if symbol else coin.upper()],
        'Crypto Name': [coin.upper()],
        'Timestamp': [current_time],
        'Timeframe': [timeframe],
        'Current Price': [last_price],
        'Prediction Value': [predicted_price if predicted_price else last_price],
        'Signal': [recommendation.get('advice', 'HOLD')],
        'Stoploss': [recommendation.get('stoploss', 0)],
        'RSI': [rsi_value],
        'Buy Time': [buy_time],
        'Sale Time': [sale_time]
    }
    
    df = pd.DataFrame(data)
    
    if EXCEL_ENHANCED:
        try:
            excel_path = output_path.replace('.csv', '.xlsx')
            
            # Check if file exists for appending
            if os.path.exists(excel_path) and os.path.getsize(excel_path) > 0:
                # Append to existing file
                wb = load_workbook(excel_path)
                ws = wb.active
                
                # Find the next serial number
                next_sl_no = ws.max_row  # Header is row 1, so next row number is the serial number
                
                # Update serial number
                df['Sl No'] = next_sl_no
                
                # Append new row
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws.append(r)
                
                print_status(f"Appended data to existing Excel file: {excel_path}", "success")
            else:
                # Create new file
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Trading Signals"
                
                # Write header
                for col_num, column_title in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_num, value=column_title)
                    ws.cell(row=1, column=col_num).font = Font(bold=True)
                
                # Write data
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws.append(r)
                
                print_status(f"Created new Excel file: {excel_path}", "success")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            return {'csv': output_path, 'xlsx': excel_path, 'enhanced': True}
            
        except Exception as e:
            print_status(f"Excel export failed: {e}, falling back to CSV", "warning")
    
    # Fallback to CSV
    csv_path = output_path
    if os.path.exists(csv_path) and os.path.getsize(csv_path) > 0:
        # Append to CSV
        df.to_csv(csv_path, mode='a', header=False, index=False)
        print_status(f"Appended data to CSV file: {csv_path}", "success")
    else:
        # Create new CSV
        df.to_csv(csv_path, index=False)
        print_status(f"Created new CSV file: {csv_path}", "success")
    
    return {'csv': csv_path, 'xlsx': None, 'enhanced': False}

# ----- NEW: Accuracy Checking Functionality -----

def load_trading_signals(file_path: str) -> pd.DataFrame:
    """Load trading signals from Excel or CSV file"""
    try:
        if file_path.endswith('.xlsx') and EXCEL_ENHANCED:
            df = pd.read_excel(file_path)
        else:
            df = pd.read_csv(file_path)
        
        # Clean column names and ensure proper data types
        df.columns = [col.strip().replace(' ', '_').lower() for col in df.columns]
        
        # Convert timestamp columns
        for col in ['timestamp', 'buy_time', 'sale_time']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
        
        # Ensure numeric columns
        numeric_cols = ['current_price', 'prediction_value', 'stoploss', 'rsi']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        print_status(f"Loaded {len(df)} trading signals from {file_path}", "success")
        return df
    except Exception as e:
        print_status(f"Error loading trading signals: {e}", "error")
        return pd.DataFrame()

def fetch_current_price(coin: str, symbol: str, vs_currency: str = 'usd') -> Optional[float]:
    """Fetch current price from CoinGecko or Binance"""
    try:
        if symbol:
            # Try Binance first for more real-time data
            try:
                url = f"https://api.binance.com/api/v3/ticker/price"
                params = {"symbol": symbol.upper()}
                r = requests.get(url, params=params, timeout=10)
                if r.status_code == 200:
                    data = r.json()
                    return float(data['price'])
            except:
                pass
        
        # Fallback to CoinGecko
        url = f"https://api.coingecko.com/api/v3/simple/price"
        params = {"ids": coin.lower(), "vs_currencies": vs_currency}
        r = requests.get(url, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
        return data[coin.lower()][vs_currency]
    except Exception as e:
        print_status(f"Error fetching current price: {e}", "error")
        return None

def calculate_accuracy_metrics(predictions_df: pd.DataFrame) -> Dict[str, Any]:
    """Calculate various accuracy metrics for predictions"""
    if predictions_df.empty:
        return {}
    
    metrics = {}
    
    # Basic accuracy calculations
    predictions_df['price_diff'] = predictions_df['actual_price'] - predictions_df['prediction_value']
    predictions_df['abs_diff'] = predictions_df['price_diff'].abs()
    predictions_df['pct_diff'] = (predictions_df['price_diff'] / predictions_df['prediction_value']) * 100
    predictions_df['abs_pct_diff'] = predictions_df['pct_diff'].abs()
    
    # Direction accuracy (whether prediction correctly guessed up/down)
    predictions_df['direction_correct'] = (
        (predictions_df['actual_price'] > predictions_df['current_price']) == 
        (predictions_df['prediction_value'] > predictions_df['current_price'])
    )
    
    # Signal accuracy (whether the trading signal was profitable)
    predictions_df['signal_profitable'] = False
    buy_signals = predictions_df['signal'].str.contains('BUY', case=False, na=False)
    predictions_df.loc[buy_signals, 'signal_profitable'] = (
        predictions_df.loc[buy_signals, 'actual_price'] > predictions_df.loc[buy_signals, 'current_price']
    )
    
    # Calculate metrics
    metrics['total_predictions'] = len(predictions_df)
    metrics['mae'] = predictions_df['abs_diff'].mean()
    metrics['mse'] = mean_squared_error(predictions_df['actual_price'], predictions_df['prediction_value'])
    metrics['rmse'] = math.sqrt(metrics['mse'])
    metrics['mape'] = predictions_df['abs_pct_diff'].mean()
    metrics['direction_accuracy'] = predictions_df['direction_correct'].mean() * 100
    metrics['signal_accuracy'] = predictions_df['signal_profitable'].mean() * 100
    
    # Profit/loss analysis
    metrics['total_profitable_signals'] = predictions_df['signal_profitable'].sum()
    metrics['profitability_rate'] = (metrics['total_profitable_signals'] / len(predictions_df)) * 100
    
    return metrics

def check_prediction_accuracy(file_path: str, vs_currency: str = 'usd') -> Dict[str, Any]:
    """
    Main function to check prediction accuracy against live data
    """
    print_header("PREDICTION ACCURACY CHECK", 1)
    
    # Load historical predictions
    predictions_df = load_trading_signals(file_path)
    if predictions_df.empty:
        return {'error': 'No data loaded'}
    
    # Filter to only include predictions that should have outcomes by now
    current_time = datetime.now()
    predictions_df = predictions_df[predictions_df['timestamp'] <= current_time]
    
    if predictions_df.empty:
        print_status("No predictions found that have reached their evaluation time", "warning")
        return {'error': 'No predictions to evaluate'}
    
    print_status(f"Evaluating {len(predictions_df)} predictions...", "info")
    
    # Fetch current prices and calculate accuracy
    results = []
    for idx, row in predictions_df.iterrows():
        coin = row.get('crypto_name', '').lower() or row.get('symbol', '').lower()
        symbol = row.get('symbol', '')
        
        if not coin:
            continue
            
        current_price = fetch_current_price(coin, symbol, vs_currency)
        if current_price is None:
            continue
            
        result = row.to_dict()
        result['actual_price'] = current_price
        result['check_timestamp'] = current_time
        results.append(result)
    
    if not results:
        print_status("No successful price fetches for accuracy calculation", "error")
        return {'error': 'Price fetch failed'}
    
    # Create results dataframe
    results_df = pd.DataFrame(results)
    
    # Calculate accuracy metrics
    metrics = calculate_accuracy_metrics(results_df)
    
    # Display results
    print_header("ACCURACY RESULTS", 2)
    
    if metrics:
        print_bullet(f"Total Predictions Evaluated: {metrics['total_predictions']}", Fore.WHITE)
        print_bullet(f"Mean Absolute Error: ${metrics['mae']:.6f}", Fore.WHITE)
        print_bullet(f"Root Mean Square Error: ${metrics['rmse']:.6f}", Fore.WHITE)
        print_bullet(f"Mean Absolute Percentage Error: {metrics['mape']:.2f}%", Fore.WHITE)
        print_bullet(f"Direction Accuracy: {metrics['direction_accuracy']:.1f}%", 
                    Fore.GREEN if metrics['direction_accuracy'] > 60 else Fore.YELLOW if metrics['direction_accuracy'] > 50 else Fore.RED)
        print_bullet(f"Signal Profitability: {metrics['signal_accuracy']:.1f}%", 
                    Fore.GREEN if metrics['signal_accuracy'] > 60 else Fore.YELLOW if metrics['signal_accuracy'] > 50 else Fore.RED)
        
        # Detailed analysis
        print_header("DETAILED ANALYSIS", 2)
        
        # Best and worst predictions
        best_pred = results_df.loc[results_df['abs_pct_diff'].idxmin()]
        worst_pred = results_df.loc[results_df['abs_pct_diff'].idxmax()]
        
        print_bullet(f"Best Prediction: {best_pred['crypto_name']} - Error: {best_pred['abs_pct_diff']:.2f}%", Fore.GREEN)
        print_bullet(f"Worst Prediction: {worst_pred['crypto_name']} - Error: {worst_pred['abs_pct_diff']:.2f}%", Fore.RED)
        
        # Timeframe analysis
        if 'timeframe' in results_df.columns:
            timeframe_stats = results_df.groupby('timeframe').agg({
                'abs_pct_diff': 'mean',
                'direction_correct': 'mean',
                'signal_profitable': 'mean'
            }).round(2)
            
            print_header("ACCURACY BY TIMEFRAME", 2)
            for timeframe, stats in timeframe_stats.iterrows():
                accuracy_color = Fore.GREEN if stats['direction_correct'] > 0.6 else Fore.YELLOW if stats['direction_correct'] > 0.5 else Fore.RED
                print_bullet(f"{timeframe}: {accuracy_color}{stats['direction_correct']*100:.1f}% accuracy, {stats['abs_pct_diff']:.2f}% avg error", Fore.WHITE)
        
        # Save detailed results
        output_file = file_path.replace('.xlsx', '_accuracy_report.xlsx').replace('.csv', '_accuracy_report.csv')
        
        if EXCEL_ENHANCED and output_file.endswith('.xlsx'):
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                results_df.to_excel(writer, sheet_name='Detailed Results', index=False)
                
                # Summary sheet
                summary_data = {
                    'Metric': list(metrics.keys()),
                    'Value': list(metrics.values())
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
        else:
            results_df.to_csv(output_file, index=False)
        
        print_status(f"Detailed accuracy report saved to: {output_file}", "success")
        
        return {
            'metrics': metrics,
            'detailed_results': results_df,
            'summary': f"Overall Accuracy: {metrics['direction_accuracy']:.1f}% | Profitability: {metrics['signal_accuracy']:.1f}%"
        }
    else:
        print_status("Could not calculate accuracy metrics", "error")
        return {'error': 'Metric calculation failed'}

# ----- Main CLI flow -----

def parse_args():
    p = argparse.ArgumentParser(description="pridict.py - Enhanced Crypto Prediction Tool with Accuracy Checking")
    p.add_argument('coin', nargs='?', help='Coin id (coingecko) or name, e.g. bitcoin, ethereum')
    p.add_argument('--tf', '--timeframe', dest='timeframe', default='1h',
                   choices=['10m', '20m', '30m', '1h', '4h', '1d', '1y'], help='timeframe to analyze')
    p.add_argument('--symbol', help='Optional exchange symbol for minute-level data (e.g. BTCUSDT) to use Binance klines')
    p.add_argument('--vs', default='usd', help='Quote currency for CoinGecko (default usd)')
    p.add_argument('--limit', type=int, default=500, help='Number of candles to fetch (where applicable)')
    p.add_argument('--news', action='store_true', help='Fetch news sentiment (requires NEWSAPI_KEY env var)')
    p.add_argument('--ollama', action='store_true', help='Generate Ollama summary (requires local Ollama running)')
    p.add_argument('--ollama-host', default=os.environ.get('OLLAMA_HOST', 'http://localhost'), help='Ollama host')
    p.add_argument('--ollama-port', type=int, default=int(os.environ.get('OLLAMA_PORT', 11500)), help='Ollama port')
    p.add_argument('--model', default='gemma3:4b', help='Ollama model name')
    p.add_argument('--out', default='pridict_output', help='Output file prefix (CSV/XLSX)')
    p.add_argument('--simple', action='store_true', help='Simplified output for beginners')
    p.add_argument('--year-prediction', action='store_true', help='Generate 1-year price prediction')
    p.add_argument('--append', action='store_true', default=True, help='Append to existing Excel file (default: True)')
    
    # New argument for accuracy checking
    p.add_argument('--check-accuracy', metavar='FILE', help='Check prediction accuracy against live data using FILE (e.g., trade.xlsx)')
    
    return p.parse_args()

def main():
    args = parse_args()
    
    # New accuracy checking mode
    if args.check_accuracy:
        if not os.path.exists(args.check_accuracy):
            print_status(f"File not found: {args.check_accuracy}", "error")
            sys.exit(1)
        
        result = check_prediction_accuracy(args.check_accuracy, args.vs)
        
        if 'error' in result:
            print_status(f"Accuracy check failed: {result['error']}", "error")
            sys.exit(1)
        else:
            print_header("ACCURACY CHECK COMPLETE ✅", 1)
            return result
    
    # Original prediction mode (requires coin argument)
    if not args.coin:
        print_status("Error: Coin argument required for prediction mode", "error")
        print_status("Use --check-accuracy FILE for accuracy checking mode", "info")
        sys.exit(1)
    
    coin = args.coin.lower()
    timeframe = args.timeframe
    vs = args.vs.lower()
    limit = args.limit
    symbol = args.symbol
    news_flag = args.news
    ollama_flag = args.ollama
    simple_mode = args.simple
    year_prediction_flag = args.year_prediction
    append_mode = args.append

    print_header(f"PRIDICT - ENHANCED CRYPTO ANALYSIS TOOL", 1)
    print_bullet(f"Coin: {Fore.CYAN}{coin.upper()}{Fore.WHITE} | Timeframe: {Fore.CYAN}{timeframe}{Fore.WHITE} | Currency: {Fore.CYAN}{vs.upper()}", Fore.WHITE, 0)
    if year_prediction_flag:
        print_bullet(f"1-Year Prediction: {Fore.GREEN}ENABLED", Fore.WHITE, 0)
    if append_mode and EXCEL_ENHANCED:
        print_bullet(f"Excel Append Mode: {Fore.GREEN}ENABLED", Fore.WHITE, 0)

    # Generate timestamp for this analysis session
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 1) Fetch price data - FIXED for 1h and 1y timeframes
    print_header("DATA COLLECTION", 2)
    df_ohlcv = None
    try:
        if symbol and timeframe in ['10m', '20m', '30m', '1h', '4h']:
            tf_map = {'10m': '10m', '20m': '20m', '30m': '30m', '1h': '1h', '4h': '4h'}
            interval = tf_map[timeframe]
            print_status(f"Fetching Binance data for {symbol} ({interval})...", "info")
            df_ohlcv = fetch_binance_klines(symbol=symbol, interval=interval, limit=limit)
        else:
            # FIXED: Proper days calculation for different timeframes
            if timeframe == '1y':
                days = 365 * 2  # 2 years of data for better 1y prediction
            elif timeframe == '1d':
                days = 365  # 1 year of daily data
            elif timeframe == '4h':
                days = 60   # 60 days for 4h data
            elif timeframe == '1h':
                days = 30   # 30 days for 1h data (was 2 days - too little)
            elif timeframe == '30m':
                days = 7    # 7 days for 30m data
            elif timeframe == '20m':
                days = 5    # 5 days for 20m data
            elif timeframe == '10m':
                days = 3    # 3 days for 10m data
            else:
                days = 30   # default fallback
            
            print_status(f"Fetching CoinGecko data for {coin} (last {days} days)...", "info")
            raw = fetch_coin_gecko_market_chart(coin_id=coin, vs_currency=vs, days=days)
            df_ohlcv = coingecko_ohlcv_from_prices(raw, timeframe)
    except Exception as e:
        print_status(f"Error fetching price data: {e}", "error")
        sys.exit(1)

    if df_ohlcv is None or df_ohlcv.empty:
        print_status("No OHLCV data available.", "error")
        sys.exit(1)

    print_status(f"Fetched {len(df_ohlcv)} price records", "success")

    # 2) Add enhanced technical indicators
    print_header("TECHNICAL ANALYSIS", 2)
    print_status("Calculating advanced technical indicators...", "info")
    df = add_technical_indicators(df_ohlcv)

    if df.empty:
        print_status("Not enough data for analysis after calculating technical indicators.", "error")
        sys.exit(1)

    # 3) ML Prediction with 1-year forecast capability
    feature_cols = ['ema8', 'ema21', 'ema50', 'rsi14', 'macd', 'macd_signal', 'macd_hist', 
                   'atr14', 'bb_width', 'volume_ratio', 'ret1', 'ret3', 'ret7']
    
    # Only use features that actually exist in the dataframe
    available_features = [col for col in feature_cols if col in df.columns]
    df_features = df.dropna(subset=available_features + ['close']).reset_index(drop=True)
    
    ml_out = None
    future_predictions = None
    
    if not df_features.empty and len(df_features) >= 50:
        try:
            print_status("Training machine learning model...", "info")
            ml_out = train_ml_predictor(df_features, available_features)
            
            if ml_out and year_prediction_flag and len(df_features) >= 100:
                print_status("Generating 1-year price prediction...", "info")
                future_predictions = predict_future_prices(
                    ml_out['model'], 
                    ml_out['last_row'], 
                    ml_out['feature_cols'], 
                    periods=365
                )
            elif year_prediction_flag:
                print_status("Skipping 1-year prediction: insufficient data", "warning")
                
        except Exception as e:
            print_status(f"ML training/prediction failed: {e}", "warning")
            ml_out = None
    else:
        print_status("Insufficient data for ML prediction", "warning")

    # Get current market state
    try:
        if ml_out and ml_out.get('last_row') is not None:
            last_row = ml_out['last_row']
        elif not df_features.empty:
            last_row = df_features.iloc[-1]
        elif not df.empty:
            last_row = df.iloc[-1]
        else:
            last_row = df_ohlcv.iloc[-1]
            
        last_price = float(last_row['close'])
        
        # Safely get other indicators with fallbacks
        last_atr = float(last_row['atr14']) if 'atr14' in last_row and is_valid_value(last_row['atr14']) else 0.02 * last_price
        last_rsi = float(last_row['rsi14']) if 'rsi14' in last_row and is_valid_value(last_row['rsi14']) else 50.0
        ema8_val = float(last_row['ema8']) if 'ema8' in last_row and is_valid_value(last_row['ema8']) else last_price
        ema21_val = float(last_row['ema21']) if 'ema21' in last_row and is_valid_value(last_row['ema21']) else last_price
        
    except (IndexError, KeyError, ValueError) as e:
        print_status(f"Error accessing market data: {e}", "error")
        last_price = float(df_ohlcv['close'].iloc[-1])
        last_atr = 0.02 * last_price
        last_rsi = 50.0
        ema8_val = last_price
        ema21_val = last_price

    # Prediction with safety check - FIXED for 1h and 1y
    predicted_next = None
    if ml_out and ml_out.get('model') is not None and ml_out.get('last_row') is not None:
        try:
            recent_feat = ml_out['last_row'][ml_out['feature_cols']].values.reshape(1, -1)
            predicted_next = float(ml_out['model'].predict(recent_feat)[0])
            if not is_valid_value(predicted_next):
                predicted_next = None
                print_status("Invalid prediction value generated", "warning")
        except Exception as e:
            print_status(f"Prediction failed: {e}", "warning")

    # 4) News sentiment
    news_result = None
    if news_flag:
        key = os.environ.get('NEWSAPI_KEY')
        print_status("Analyzing news sentiment...", "info")
        news_result = fetch_news_sentiment(coin, key)
        if not news_result.get('available'):
            print_status(f"News unavailable: {news_result.get('reason')}", "warning")

    # 5) Whale analysis
    whale_res = None
    if symbol:
        try:
            print_status("Analyzing whale transactions...", "info")
            trades = fetch_binance_recent_trades(symbol=symbol, limit=200)
            whale_res = analyze_whales_from_trades(trades)
        except Exception as e:
            whale_res = {'available': False, 'reason': str(e)}

    # 6) Ollama summary
    ollama_summary = None
    if ollama_flag:
        print_status("Generating AI analysis...", "info")
        prompt = f"Provide a concise trading summary for {coin} at ${last_price:.2f} based on {timeframe} timeframe. Key indicators: RSI={last_rsi:.1f}, Trend analysis needed. Give simple advice."
        try:
            out = generate_ollama_summary(prompt, host=args.ollama_host, port=args.ollama_port, model=args.model)
            if out.get('ok'):
                ollama_summary = out.get('text')
            else:
                ollama_summary = f"AI analysis unavailable: {out.get('error')}"
        except Exception as e:
            ollama_summary = f"AI analysis failed: {e}"

    # 7) Enhanced recommendation with timeframe-based sale time
    recommendation = recommend_trade(last_price, predicted_next, last_atr, last_rsi, timeframe)
    rsi_strength, rsi_color = get_rsi_strength(last_rsi)
    trend_strength, trend_color = get_trend_strength(ema8_val, ema21_val, last_price)

    # 8) OUTPUT
    print_header("📊 MARKET ANALYSIS RESULTS", 1)
    
    # Price Section
    print_header("Price Information", 2)
    print_bullet(f"Current Price: {Fore.MAGENTA}${last_price:.8f}", Fore.WHITE)
    if predicted_next:
        change_color = Fore.GREEN if predicted_next > last_price else Fore.RED
        change_icon = "📈" if predicted_next > last_price else "📉"
        print_bullet(f"Predicted Price: {change_color}${predicted_next:.8f} {change_icon}", Fore.WHITE)
        print_bullet(f"Expected Change: {change_color}{recommendation['diff_pct']:.3f}%", Fore.WHITE)
    
    # Technical Indicators
    print_header("Technical Indicators", 2)
    print_bullet(f"RSI (14): {rsi_color}{last_rsi:.1f} - {rsi_strength}", Fore.WHITE)
    print_bullet(f"Trend: {trend_color}{trend_strength}", Fore.WHITE)
    print_bullet(f"Volatility (ATR): ${last_atr:.8f}", Fore.WHITE)
    print_bullet(f"EMA8: ${ema8_val:.8f} | EMA21: ${ema21_val:.8f}", Fore.WHITE)
    
    # Trading Recommendation with Sale Time
    print_header("TRADING RECOMMENDATION", 2)
    advice_color = Fore.GREEN if 'BUY' in recommendation['advice'] else Fore.RED if 'SELL' in recommendation['advice'] else Fore.YELLOW
    print_bullet(f"Action: {advice_color}{recommendation['advice']} {Style.BRIGHT}", Fore.WHITE)
    print_bullet(f"Confidence: {recommendation['confidence']}", Fore.WHITE)
    
    # Sale Time Information
    if recommendation.get('sale_time'):
        sale_time = recommendation['sale_time']
        print_bullet(f"Recommended Sale Time: {Fore.CYAN}{sale_time.strftime('%Y-%m-%d %H:%M:%S')}", Fore.WHITE)
        time_until_sale = sale_time - datetime.now()
        hours_until = time_until_sale.total_seconds() / 3600
        if hours_until > 24:
            days_until = hours_until / 24
            print_bullet(f"Time Until Sale: {Fore.CYAN}{days_until:.1f} days", Fore.WHITE)
        else:
            print_bullet(f"Time Until Sale: {Fore.CYAN}{hours_until:.1f} hours", Fore.WHITE)
    
    # Risk Management
    if recommendation['stoploss']:
        print_header("Risk Management", 2)
        print_bullet(f"Stop Loss: ${recommendation['stoploss']:.8f}", Fore.WHITE)
        print_bullet(f"Take Profit: ${recommendation['take_profit']:.8f}", Fore.WHITE)
        if recommendation['risk_reward_ratio'] and is_valid_value(recommendation['risk_reward_ratio']):
            rr_color = Fore.GREEN if recommendation['risk_reward_ratio'] > 2 else Fore.YELLOW if recommendation['risk_reward_ratio'] > 1 else Fore.RED
            print_bullet(f"Risk/Reward Ratio: {rr_color}{recommendation['risk_reward_ratio']:.2f}", Fore.WHITE)
    
    # Market Factors
    print_header("Market Factors", 2)
    
    # Whale Activity
    if whale_res and whale_res.get('available'):
        whales = whale_res.get('whales', [])
        if whales:
            whale_count = len(whales)
            if whale_count > 10:
                whale_color = Fore.RED
                whale_text = f"Very High ({whale_count} large trades detected)"
            elif whale_count > 5:
                whale_color = Fore.YELLOW
                whale_text = f"High ({whale_count} large trades detected)"
            else:
                whale_color = Fore.GREEN
                whale_text = f"Moderate ({whale_count} large trades detected)"
            print_bullet(f"Whale Activity: {whale_color}{whale_text}", Fore.WHITE)
        else:
            print_bullet("Whale Activity: Low (no significant large trades)", Fore.WHITE)
    elif symbol:
        print_bullet("Whale Activity: Data unavailable (symbol required)", Fore.WHITE)
    else:
        print_bullet("Whale Activity: Analysis skipped (no symbol provided)", Fore.WHITE)
    
    # News Sentiment
    if news_result and news_result.get('available'):
        sentiment_score = news_result.get('avg_score', 0)
        sentiment_color = Fore.GREEN if sentiment_score > 0.5 else Fore.RED if sentiment_score < -0.5 else Fore.YELLOW
        sentiment_text = "Positive" if sentiment_score > 0.5 else "Negative" if sentiment_score < -0.5 else "Neutral"
        print_bullet(f"News Sentiment: {sentiment_color}{sentiment_text} (score: {sentiment_score:.2f})", Fore.WHITE)
    elif news_flag:
        print_bullet("News Sentiment: Analysis unavailable (check NEWSAPI_KEY)", Fore.WHITE)
    else:
        print_bullet("News Sentiment: Analysis skipped (use --news to enable)", Fore.WHITE)
    
    # ML Model Performance
    if ml_out and ml_out.get('mae'):
        mae = ml_out['mae']
        mae_pct = (mae / last_price) * 100
        mae_color = Fore.GREEN if mae_pct < 0.5 else Fore.YELLOW if mae_pct < 1.0 else Fore.RED
        print_bullet(f"ML Model Accuracy: {mae_color}MAE {mae_pct:.3f}%", Fore.WHITE)
    
    # 9) Simplified Excel Export with only required columns including Sale Time
    print_header("DATA EXPORT", 2)
    
    output_file = f"{args.out}_trading_signals.xlsx"
    export_result = create_simplified_export(
        coin=coin,
        symbol=symbol,
        timeframe=timeframe,
        last_price=last_price,
        predicted_price=predicted_next,
        recommendation=recommendation,
        rsi_value=last_rsi,
        output_path=output_file,
        timestamp_str=timestamp_str
    )
    
    if export_result['enhanced']:
        print_status(f"Simplified trading data exported to: {export_result['xlsx']}", "success")
    else:
        print_status(f"Simplified trading data exported to: {export_result['csv']}", "success")

    # Display 1-year prediction if available
    if future_predictions is not None and not future_predictions.empty:
        print_header("1-YEAR PRICE PREDICTION", 2)
        current_price = last_price
        predicted_1y = future_predictions['predicted_price'].iloc[-1]
        change_pct = ((predicted_1y - current_price) / current_price) * 100
        
        print_bullet(f"Current Price: ${current_price:.8f}", Fore.WHITE)
        print_bullet(f"Predicted 1-Year Price: ${predicted_1y:.8f}", Fore.GREEN if change_pct > 0 else Fore.RED)
        print_bullet(f"Expected Change: {change_pct:.2f}%", Fore.GREEN if change_pct > 0 else Fore.RED)
        
        # Save prediction to separate file
        pred_csv = f"{args.out}_1y_prediction.csv"
        future_predictions.to_csv(pred_csv, index=False)
        print_status(f"1-year prediction saved to: {pred_csv}", "success")

    # Simple Mode Summary
    if simple_mode:
        print_header("🚀 QUICK SUMMARY", 1)
        print_bullet(f"Action: {advice_color}{recommendation['advice']}", Fore.WHITE, 0)
        print_bullet(f"Price: ${last_price:.2f} | Trend: {trend_strength}", Fore.WHITE, 0)
        print_bullet(f"Confidence: {recommendation['confidence']}", Fore.WHITE, 0)
        if recommendation.get('sale_time'):
            print_bullet(f"Sale Time: {recommendation['sale_time'].strftime('%Y-%m-%d %H:%M')}", Fore.WHITE, 0)
    
    # Ollama AI Summary
    if ollama_summary and not simple_mode:
        print_header("AI ANALYSIS", 2)
        print(f"{Fore.CYAN}{ollama_summary}")
    
    # Final Disclaimer
    print_header("⚠️  DISCLAIMER", 2)
    print_bullet("This analysis is for educational purposes only", Fore.YELLOW)
    print_bullet("Always do your own research before trading", Fore.YELLOW)
    print_bullet("Never invest more than you can afford to lose", Fore.YELLOW)
    
    print_header("ANALYSIS COMPLETE ✅", 1)
    
    return {
        'last_price': last_price,
        'predicted_next': predicted_next,
        'future_predictions': future_predictions,
        'recommendation': recommendation,
        'export_paths': export_result,
        'ml_mae': (ml_out['mae'] if ml_out else None),
        'timestamp': timestamp_str
    }

if __name__ == "__main__":
    res = main()